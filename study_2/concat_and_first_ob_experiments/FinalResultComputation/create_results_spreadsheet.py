import pandas as pd
import os
import glob
from openpyxl.styles import Color, PatternFill, Font, Border, Side


def highlight_cells_for_larger_values(sheet_name, writer, column_name, base_value):
    yellow_fill = PatternFill(start_color='00FFFF00',
                              end_color='00FFFF00',
                              fill_type='solid')

    for cell in writer.sheets[sheet_name][column_name]:
        if isinstance(cell.value, str):
            continue
        if cell.value > base_value:
            cell.fill = yellow_fill


def highlight_cells_for_smaller_values(sheet_name,  writer, column_name, base_value):
    yellow_fill = PatternFill(start_color='00FFFF00',
                              end_color='00FFFF00',
                              fill_type='solid')

    for cell in writer.sheets[sheet_name][column_name]:
        if isinstance(cell.value, str):
            continue
        if cell.value < base_value:
            cell.fill = yellow_fill


def set_border(ws, staring_range, ending_range):
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    rows = ws[staring_range:ending_range]
    for row in rows:
        for cell in row:
            cell.border = border


def modify_df_and_save_to_excel(df, writer, sheet_name):
    df.style.set_properties(**{'text-align': 'center'}).to_excel(writer, sheet_name=sheet_name, index=False, float_format="%.3f")
    worksheet = writer.sheets[sheet_name]

    worksheet.freeze_panes = 'A2'

    baseline_row = df.iloc[63]
    # print(baseline_row)
    avg_best_ranks = baseline_row['Avg Best Ranks'].round(3)
    highlight_cells_for_smaller_values(sheet_name, writer, 'G', avg_best_ranks)
    n_identified_bugs = baseline_row['Number of Identified Bugs']
    highlight_cells_for_larger_values(sheet_name, writer, 'H', n_identified_bugs)
    hit_1 = baseline_row['Hit@1'].round(3)
    highlight_cells_for_larger_values(sheet_name, writer, 'J', hit_1)
    hit_2 = baseline_row['Hit@2'].round(3)
    highlight_cells_for_larger_values(sheet_name, writer, 'K', hit_2)
    hit_3 = baseline_row['Hit@3'].round(3)
    highlight_cells_for_larger_values(sheet_name, writer, 'L', hit_3)
    hit_4 = baseline_row['Hit@4'].round(3)
    highlight_cells_for_larger_values(sheet_name, writer, 'M', hit_4)
    hit_5 = baseline_row['Hit@5'].round(3)
    highlight_cells_for_larger_values(sheet_name, writer, 'N', hit_5)
    hit_10 = baseline_row['Hit@10'].round(3)
    highlight_cells_for_larger_values(sheet_name, writer, 'O', hit_10)
    n_hits_10 = baseline_row['Number of hits@10']
    highlight_cells_for_larger_values(sheet_name, writer, 'P', n_hits_10)
    avg_hits_10 = baseline_row['Avg Hit@10'].round(3)
    highlight_cells_for_larger_values(sheet_name, writer, 'Q', avg_hits_10)
    hit_15 = baseline_row['Hit@15'].round(3)
    highlight_cells_for_larger_values(sheet_name, writer, 'R', hit_15)
    hit_20 = baseline_row['Hit@20'].round(3)
    highlight_cells_for_larger_values(sheet_name, writer, 'S', hit_20)
    avg_hits_20 = baseline_row['Avg Hit@20'].round(3)
    highlight_cells_for_larger_values(sheet_name, writer, 'T', avg_hits_20)
    mrr = baseline_row['MRR'].round(3)
    highlight_cells_for_larger_values(sheet_name, writer, 'U', mrr)
    map = baseline_row['MAP'].round(3)
    highlight_cells_for_larger_values(sheet_name, writer, 'V', map)

    # Set the border
    set_border(worksheet, 'A1', 'W71')
    writer._save()


def create_spreadsheet(input_file_path, sheet_name, output_file_path):
    # Read the csv file
    df = pd.read_csv(input_file_path)

    # Write the dataframe to the excel file
    if os.path.exists(output_file_path):
        with pd.ExcelWriter(output_file_path, mode="a", engine="openpyxl") as writer:
            modify_df_and_save_to_excel(df, writer, sheet_name)
    else:
        with pd.ExcelWriter(output_file_path, mode="w", engine="openpyxl") as writer:
            modify_df_and_save_to_excel(df, writer, sheet_name)


if __name__ == "__main__":
    results_folder_path = './FinalResults'
    lucene_summary_file_path = './ResultsSummary/BL-Exp-Results-LUCENE.xlsx'
    bug_locator_summary_file_path = './ResultsSummary/BL-Exp-Results-BugLocator.xlsx'

    # Remove the output file if it exists
    if os.path.exists(lucene_summary_file_path):
        os.remove(lucene_summary_file_path)
    if os.path.exists(bug_locator_summary_file_path):
        os.remove(bug_locator_summary_file_path)

    files = glob.glob(os.path.join(results_folder_path, "*.csv"))
    files.sort()

    # Iterate over all the files in the results folder
    for file in files:
        # Get the sheet name from the file name
        exp_name = file.split("/")[-1].replace(".csv", "")
        if exp_name.startswith("LUCENE"):
            print(exp_name)
            sheet_name = exp_name.replace("LUCENE-", "")
            create_spreadsheet(file, sheet_name, lucene_summary_file_path)
        elif exp_name.startswith("BugLocator"):
            print(exp_name)
            sheet_name = exp_name.replace("BugLocator-", "")
            create_spreadsheet(file, sheet_name, bug_locator_summary_file_path)

